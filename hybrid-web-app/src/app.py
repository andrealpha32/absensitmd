from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timezone, timedelta
from models import db, Student, Attendance, AttendanceSummary, ProfileChangeLog
from config import Config
from sqlalchemy import case
import logging
import os
import calendar
import schedule
import time
import base64
import io
import csv
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except ImportError:
    from pytz import timezone as ZoneInfo  # fallback jika pakai pytz

app = Flask(__name__)
app.config.from_object(Config)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')

# Define allowed file extensions
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

# Create uploads directory if it doesn't exist
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

db.init_app(app)

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Add hardcoded admin credentials
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def home():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        # Log the login attempt
        logger.debug(f"Login attempt for email: {email}")
        
        if not email or not password:
            flash('Mohon isi email dan password', 'danger')
            return render_template('login.html')
            
        # Find the student and log the result
        student = Student.query.filter_by(email=email).first()
        if student:
            logger.debug(f"Found student: {student.name}, checking password")
            # Check if passwords match exactly
            if student.password == password:
                session['logged_in'] = True
                session['student_id'] = student.id
                session['student_name'] = student.name  # Add name to session for convenience
                logger.debug(f"Login successful for student: {student.name}")
                flash('Login berhasil!', 'success')
                return redirect(url_for('dashboard'))
            else:
                logger.debug("Password mismatch")
                flash('Password salah!', 'danger')
        else:
            logger.debug("Student not found with provided email")
            flash('Email tidak terdaftar!', 'danger')
            
    return render_template('login.html')

def get_today_attendance(student_id):
    student = db.session.get(Student, student_id)  # Updated from query.get()
    if not student:
        return None
    
    return Attendance.query.filter_by(
        student_id=student.id,
        date=datetime.now(timezone.utc).date()  # Updated to use timezone-aware datetime
    ).first()

@app.route('/dashboard')
def dashboard():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    student = db.session.get(Student, session['student_id'])  # Updated from query.get()
    
    # Get current month calendar data
    current_date = datetime.now(timezone.utc)  # Using timezone-aware datetime
    cal = calendar.monthcalendar(current_date.year, current_date.month)
    today = current_date.date()
    
    # Get attendance summaries for current month
    summaries = AttendanceSummary.query.filter(
        db.extract('month', AttendanceSummary.date) == today.month,
        db.extract('year', AttendanceSummary.date) == today.year
    ).all()
    
    # Format calendar data
    calendar_data = []
    for week in cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'date': ''})
                continue
                
            date = datetime(today.year, today.month, day, tzinfo=timezone.utc).date()
            summary = next((s for s in summaries if s.date == date), None)
            
            week_data.append({
                'date': day,
                'today': date == today,
                'past': date < today,
                'summary': summary
            })
        calendar_data.append(week_data)
    
    return render_template('dashboard.html',
                         student=student,
                         attendance=get_today_attendance(session['student_id']),
                         calendar_data=calendar_data,
                         current_month=current_date.strftime('%B %Y'))

@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    if not session.get('logged_in'):
        flash('Please login first', 'warning')
        return redirect(url_for('login'))
    
    student_id = session.get('student_id')
    if not student_id:
        flash('Session invalid, please login again', 'warning')
        return redirect(url_for('logout'))
    
    student = Student.query.get(student_id)
    if not student:
        flash('Student not found, please login again', 'warning')
        return redirect(url_for('logout'))
    
    # Gunakan zona waktu Asia/Jakarta
    WIB = ZoneInfo('Asia/Jakarta')
    now_wib = datetime.now(WIB)
    today = now_wib.date()
    
    # Check if student has already attended today
    existing_attendance = Attendance.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if existing_attendance:
        flash('Anda sudah melakukan absensi hari ini!', 'warning')
        return redirect(url_for('dashboard'))
    
    # Check time range (08:00 - 17:00 WIB)
    current_time = now_wib.time()
    start_time = datetime.strptime('08:00', '%H:%M').time()
    end_time = datetime.strptime('17:00', '%H:%M').time()
    if not (start_time <= current_time <= end_time):
        flash('Absensi hanya dapat dilakukan antara pukul 08:00 sampai 17:00 WIB!', 'warning')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            activity = request.form.get('activity')
            attendance_type = request.form.get('attendance_type')
            photo_data = request.form.get('photo')  # Base64 image data
            
            # Validation
            if not activity or not attendance_type:
                return jsonify({
                    'success': False, 
                    'message': 'Mohon isi semua field yang diperlukan!'
                }), 400
            
            # Handle photo upload
            photo_path = None
            if photo_data and photo_data.startswith('data:image/'):
                try:
                    # Extract base64 data
                    header, data = photo_data.split(',', 1)
                    image_data = base64.b64decode(data)
                    
                    # Generate filename
                    filename = f"{student_id}_{now_wib.strftime('%Y%m%d_%H%M%S')}.jpg"
                    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    
                    # Save image
                    with open(save_path, 'wb') as f:
                        f.write(image_data)
                    
                    photo_path = f"uploads/{filename}"
                    
                except Exception as e:
                    logger.error(f"Error saving photo: {str(e)}")
                    return jsonify({
                        'success': False, 
                        'message': 'Gagal menyimpan foto!'
                    }), 500
            
            # Create attendance record
            attendance = Attendance(
                student_id=student_id,
                date=today,
                time_in=now_wib.time(),
                activity=activity,
                attendance_type=attendance_type,
                photo_path=photo_path,
                status=attendance_type
            )
            
            db.session.add(attendance)
            db.session.commit()
            
            logger.info(f"Attendance recorded for student {student_id} on {today}")
            
            return jsonify({
                'success': True, 
                'message': 'Absensi berhasil dicatat!'
            })
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error recording attendance: {str(e)}")
            return jsonify({
                'success': False, 
                'message': 'Terjadi kesalahan saat mencatat absensi. Silakan coba lagi.'
            }), 500

    return render_template('attendance.html')

@app.route('/history')
def history():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    student = Student.query.get(session['student_id'])
    attendances = Attendance.query.filter_by(student_id=student.id).order_by(Attendance.date.desc()).all()
    return render_template('history.html', attendances=attendances)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Use hardcoded credentials instead of database
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            flash('Login berhasil!', 'success')
            return redirect(url_for('admin_dashboard'))
        
        flash('Username atau password salah!', 'danger')
    
    return render_template('admin/login.html')

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    # Add year and month to template variables
    now = datetime.now()
    today = now.date()
    year = int(request.args.get('year', now.year))
    month = int(request.args.get('month', now.month))
    
    # Get calendar data with detailed student attendance for all days
    cal = calendar.monthcalendar(year, month)
    
    # Get all attendance records for the month
    month_attendance = db.session.query(
        Attendance.date,
        Attendance.status,
        Student
    ).join(Student).filter(
        db.extract('year', Attendance.date) == year,
        db.extract('month', Attendance.date) == month
    ).all()

    # Group attendance by date
    attendance_by_date = {}
    for attendance in month_attendance:
        date = attendance.date
        if date not in attendance_by_date:
            attendance_by_date[date] = {
                'hadir': 0,
                'sakit': 0,
                'izin': 0,
                'students': []
            }
        attendance_by_date[date][attendance.status] += 1
        attendance_by_date[date]['students'].append({
            'name': attendance.Student.name,
            'status': attendance.status
        })

    # Format calendar data
    calendar_data = []
    for week in cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'date': ''})
                continue
                
            date = datetime(year, month, day).date()
            attendance_data = attendance_by_date.get(date, {
                'hadir': 0,
                'sakit': 0,
                'izin': 0,
                'students': []
            })
            day_data = {
                'date': day,
                'today': date == today,
                'past': date < today,
                'attendance': attendance_data if attendance_data['hadir'] > 0 or attendance_data['sakit'] > 0 or attendance_data['izin'] > 0 else None
            }
            week_data.append(day_data)
        calendar_data.append(week_data)

    # Get all students and their attendance for today
    today_attendance = db.session.query(
        Student,
        Attendance
    ).outerjoin(
        Attendance,
        db.and_(
            Attendance.student_id == Student.id,
            Attendance.date == today
        )
    ).order_by(Student.name).all()  # Order by student name for consistent display

    # Format today's students data
    students_today = []
    for student, attendance in today_attendance:
        student_data = {
            'id': student.id,
            'name': student.name,
            'email': student.email,
            'company': 'Trimediatama',
            'attendance_status': attendance.status if attendance else 'Belum Absen',
            'time_in': attendance.time_in.strftime('%H:%M') if attendance and attendance.time_in else None,
            'activity': attendance.activity if attendance else None,
            'photo_path': attendance.photo_path if attendance and attendance.photo_path else None
        }
        students_today.append(student_data)

    # Get basic statistics
    statistics = {
        'total_students': Student.query.count(),
        'today_attendance': Attendance.query.filter_by(date=today).count(),
        'total_attendance': len(month_attendance)
    }

    # Hitung prev/next month & year
    if month == 1:
        prev_month = 12
        prev_year = year - 1
    else:
        prev_month = month - 1
        prev_year = year
    if month == 12:
        next_month = 1
        next_year = year + 1
    else:
        next_month = month + 1
        next_year = year

    return render_template('admin/dashboard.html',
                         calendar_data=calendar_data,
                         current_month=calendar.month_name[month] + ' ' + str(year),
                         year=year,
                         month=month,
                         statistics=statistics,
                         students=students_today,
                         total_students=len(students_today),
                         calendar=calendar,
                         prev_month=prev_month,
                         prev_year=prev_year,
                         next_month=next_month,
                         next_year=next_year)

@app.route('/admin/student/<int:student_id>')
def admin_student_detail(student_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    student = Student.query.get_or_404(student_id)
    attendances = Attendance.query.filter_by(student_id=student_id).order_by(Attendance.date.desc()).all()
    return render_template('admin/student_detail.html', student=student, attendances=attendances)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/attendance/<date>')
def admin_get_day_attendance(date):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        
        # Query all attendances for the given date with student information
        attendances = db.session.query(
            Attendance,
            Student
        ).join(Student).filter(
            Attendance.date == date_obj
        ).all()
        
        # Format the response
        attendance_data = []
        for att, student in attendances:
            attendance_data.append({
                'student_id': student.id,
                'student_name': student.name,
                'status': att.status or 'Belum Absen',
                'time_in': att.time_in.strftime('%H:%M') if att.time_in else '-',
                'activity': att.activity or '-',
                'photo_path': att.photo_path if att.photo_path else None
            })
        
        return jsonify(attendance_data)
    except Exception as e:
        app.logger.error(f"Error fetching attendance data: {str(e)}")
        return jsonify([])  # Return empty array instead of error

@app.route('/api/attendance/<int:year>/<int:month>/<int:day>')
def get_detailed_attendance(year, month, day):  # Renamed from get_day_attendance
    try:
        # Convert to datetime
        selected_date = datetime(year, month, day).date()
        
        # Query all attendance records for the selected date
        attendances = Attendance.query.filter(
            db.func.date(Attendance.date) == selected_date
        ).join(Student).all()
        
        # Format the attendance data
        attendance_data = []
        for attendance in attendances:
            attendance_data.append({
                'id': attendance.id,
                'student_name': attendance.student.name,
                'status': attendance.attendance_type,
                'time_in': attendance.time_in.strftime('%H:%M') if attendance.time_in else None,
                'activity': attendance.activity,
                'photo_path': attendance.photo_path if attendance.photo_path else None
            })
        
        return jsonify(attendance_data)
    except Exception as e:
        app.logger.error(f"Error fetching attendance data: {str(e)}")
        return jsonify({'error': 'Failed to fetch attendance data'}), 500

# Test endpoint to check student data
@app.route('/test_student/<email>')
def test_student(email):
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
        
    student = Student.query.filter_by(email=email).first()
    if student:
        return jsonify({
            'id': student.id,
            'name': student.name,
            'email': student.email,
            'password_length': len(student.password) if student.password else 0
        })
    return jsonify({'error': 'Student not found'}), 404

# Function to ensure attendance data is properly saved
def save_attendance_summary(date):
    # Check if summary exists
    summary = AttendanceSummary.query.filter_by(date=date).first()
    if not summary:
        # Get attendance counts for the date
        attendance_counts = db.session.query(
            db.func.sum(db.case((Attendance.status == 'hadir', 1), else_=0)).label('present'),
            db.func.sum(db.case((Attendance.status == 'sakit', 1), else_=0)).label('sick'),
            db.func.sum(db.case((Attendance.status == 'izin', 1), else_=0)).label('excused')
        ).filter(Attendance.date == date).first()

        # Create new summary
        summary = AttendanceSummary(
            date=date,
            present_count=attendance_counts.present or 0,
            sick_count=attendance_counts.sick or 0,
            excused_count=attendance_counts.excused or 0
        )
        db.session.add(summary)
        db.session.commit()
        logger.info(f"Created attendance summary for date: {date}")

# Automatically save summary when recording attendance
@app.after_request
def after_request(response):
    if request.endpoint == 'attendance' and request.method == 'POST':
        today = datetime.now().date()
        save_attendance_summary(today)
    return response

# Add this function to reset daily attendance counts
def reset_daily_attendance_counts():
    try:
        with app.app_context():
            # Reset all students' daily attendance counts
            students = Student.query.all()
            for student in students:
                student.attendance_count = {
                    'present': 0,
                    'late': 0
                }
            db.session.commit()
            logging.info("Daily attendance counts have been reset")
    except Exception as e:
        logging.error(f"Error resetting attendance counts: {str(e)}")

# Add this to your main section
# if __name__ == '__main__':
#     # Schedule the daily reset at midnight
#     schedule.every().day.at("00:00").do(reset_daily_attendance_counts)
#     
#     # Run the scheduler in a separate thread
#     import threading
#     def run_scheduler():
#         while True:
#             schedule.run_pending()
#             time.sleep(60)
#     
#     scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
#     scheduler_thread.start()
#     
#     with app.app_context():
#         db.create_all()
#     app.run(debug=True, host='0.0.0.0', port=5000)

@app.route('/api/attendance-stats')
def attendance_stats():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401

    # Get monthly attendance statistics
    current_date = datetime.now()
    current_month = current_date.month
    current_year = current_date.year

    # Query for monthly attendance
    monthly_attendance = db.session.query(
        Student.name,
        db.func.count(Attendance.id).label('attendance_count'),
        db.func.sum(case([(Attendance.status == 'Terlambat', 1)], else_=0)).label('late_count')
    ).join(Attendance).filter(
        db.extract('month', Attendance.date) == current_month,
        db.extract('year', Attendance.date) == current_year
    ).group_by(Student.id, Student.name).all()

    # Get attendance statistics for all students
    all_students = Student.query.all()
    total_days = db.session.query(db.func.count(db.distinct(Attendance.date))).scalar() or 1
    
    student_stats = []
    for student in all_students:
        # Count total attendances
        attendance_count = db.session.query(Attendance).filter(
            Attendance.student_id == student.id,
            db.extract('month', Attendance.date) == current_month,
            db.extract('year', Attendance.date) == current_year
        ).count()
        
        # Calculate absent rate
        absent_days = total_days - attendance_count
        
        student_stats.append({
            'name': student.name,
            'attendance_count': attendance_count,
            'absent_days': absent_days
        })
    
    # Sort for most present and most absent students
    most_present = sorted(student_stats, key=lambda x: x['attendance_count'], reverse=True)[:5]
    most_absent = sorted(student_stats, key=lambda x: x['absent_days'], reverse=True)[:5]

    # Get weekly attendance data
    weekly_data = db.session.query(
        db.func.date(Attendance.date).label('date'),
        db.func.count(case([(Attendance.status == 'Hadir', 1)], else_=0)).label('present_count'),
        db.func.count(case([(Attendance.status == 'Terlambat', 1)], else_=0)).label('late_count')
    ).filter(
        Attendance.date >= datetime.now() - timedelta(days=7)
    ).group_by(db.func.date(Attendance.date)).all()

    weekly_stats = {
        'labels': [],
        'present': [],
        'late': []
    }

    for day in weekly_data:
        weekly_stats['labels'].append(day.date.strftime('%Y-%m-%d'))
        weekly_stats['present'].append(day.present_count)
        weekly_stats['late'].append(day.late_count)

    return jsonify({
        'weekly_stats': weekly_stats,
        'most_present': most_present,
        'most_absent': most_absent
    })

@app.route('/api/admin/attendance-stats')
def admin_attendance_stats():
    if not session.get('is_admin'):
        return jsonify({'error': 'Unauthorized'}), 401

    # Get monthly attendance statistics
    current_date = datetime.now()
    current_month = current_date.month
    current_year = current_date.year

    # Calculate overall statistics
    total_students = Student.query.count()
    total_days = db.session.query(db.func.count(db.distinct(Attendance.date))).scalar() or 1
    
    # Get daily attendance counts for the last 30 days
    daily_stats = db.session.query(
        db.func.date(Attendance.date).label('date'),
        db.func.count(Attendance.id).label('total_present'),
        db.func.count(case([(Attendance.status == 'Terlambat', 1)], else_=0)).label('total_late')
    ).filter(
        Attendance.date >= datetime.now() - timedelta(days=30)
    ).group_by(db.func.date(Attendance.date)).all()

    # Get student attendance statistics
    student_stats = db.session.query(
        Student.name,
        Student.email,
        db.func.count(Attendance.id).label('attendance_count')
    ).outerjoin(Attendance).group_by(Student.id).all()

    # Calculate attendance rates and prepare statistics
    attendance_stats = []
    for stat in student_stats:
        attendance_rate = (stat.attendance_count / total_days) * 100
        attendance_stats.append({
            'name': stat.name,
            'email': stat.email,
            'attendance_rate': round(attendance_rate, 2),
            'present_days': stat.attendance_count,
            'absent_days': total_days - stat.attendance_count
        })

    # Sort by attendance rate
    most_present = sorted(attendance_stats, key=lambda x: x['attendance_rate'], reverse=True)[:5]
    most_absent = sorted(attendance_stats, key=lambda x: x['absent_days'], reverse=True)[:5]

    # Prepare daily statistics for the chart
    daily_data = {
        'labels': [],
        'present': [],
        'absent': []
    }

    for day in daily_stats:
        daily_data['labels'].append(day.date.strftime('%Y-%m-%d'))
        daily_data['present'].append(day.total_present)
        daily_data['absent'].append(total_students - day.total_present)

    return jsonify({
        'overall_stats': {
            'total_students': total_students,
            'total_days': total_days,
            'average_attendance_rate': round(sum(s['attendance_rate'] for s in attendance_stats) / len(attendance_stats), 2)
        },
        'daily_stats': daily_data,
        'most_present': most_present,
        'most_absent': most_absent
    })
    
    # Profile Routes - Tambahkan ke app.py setelah route yang sudah ada

@app.route('/profile')
def profile():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    student = Student.query.get(session['student_id'])
    if not student:
        flash('Student not found', 'error')
        return redirect(url_for('logout'))
    
    # Calculate profile completion percentage
    profile_fields = [
        student.phone, student.address, student.bio, student.birth_date,
        student.gender, student.education_level, student.school_name,
        student.internship_start_date, student.internship_end_date
    ]
    completed_fields = sum(1 for field in profile_fields if field)
    completion_percentage = int((completed_fields / len(profile_fields)) * 100)
    
    # Get recent attendance statistics
    recent_attendances = Attendance.query.filter_by(student_id=student.id).order_by(Attendance.date.desc()).limit(5).all()
    
    # Get profile change history
    recent_changes = ProfileChangeLog.query.filter_by(student_id=student.id).order_by(ProfileChangeLog.changed_at.desc()).limit(5).all()
    
    return render_template('profile.html', 
                         student=student, 
                         completion_percentage=completion_percentage,
                         recent_attendances=recent_attendances,
                         recent_changes=recent_changes)

@app.route('/profile/edit', methods=['GET', 'POST'])
def edit_profile():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    student = Student.query.get(session['student_id'])
    if not student:
        flash('Student not found', 'error')
        return redirect(url_for('logout'))
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form.get('name')
            phone = request.form.get('phone')
            address = request.form.get('address')
            bio = request.form.get('bio')
            birth_date = request.form.get('birth_date')
            gender = request.form.get('gender')
            education_level = request.form.get('education_level')
            school_name = request.form.get('school_name')
            internship_start_date = request.form.get('internship_start_date')
            internship_end_date = request.form.get('internship_end_date')
            
            # Track changes for audit log
            changes = []
            if student.name != name:
                changes.append(('name', student.name, name))
                student.name = name
            if student.phone != phone:
                changes.append(('phone', student.phone, phone))
                student.phone = phone
            if student.address != address:
                changes.append(('address', student.address, address))
                student.address = address
            if student.bio != bio:
                changes.append(('bio', student.bio, bio))
                student.bio = bio
            if student.gender != gender:
                changes.append(('gender', student.gender, gender))
                student.gender = gender
            if student.education_level != education_level:
                changes.append(('education_level', student.education_level, education_level))
                student.education_level = education_level
            if student.school_name != school_name:
                changes.append(('school_name', student.school_name, school_name))
                student.school_name = school_name
            
            # Handle date fields
            if birth_date:
                birth_date_obj = datetime.strptime(birth_date, '%Y-%m-%d').date()
                if student.birth_date != birth_date_obj:
                    changes.append(('birth_date', str(student.birth_date), birth_date))
                    student.birth_date = birth_date_obj
            
            if internship_start_date:
                start_date_obj = datetime.strptime(internship_start_date, '%Y-%m-%d').date()
                if student.internship_start_date != start_date_obj:
                    changes.append(('internship_start_date', str(student.internship_start_date), internship_start_date))
                    student.internship_start_date = start_date_obj
            
            if internship_end_date:
                end_date_obj = datetime.strptime(internship_end_date, '%Y-%m-%d').date()
                if student.internship_end_date != end_date_obj:
                    changes.append(('internship_end_date', str(student.internship_end_date), internship_end_date))
                    student.internship_end_date = end_date_obj
            
            # Update timestamp
            student.updated_at = datetime.utcnow()
            
            # Log changes
            for field, old_val, new_val in changes:
                log_entry = ProfileChangeLog(
                    student_id=student.id,
                    field_changed=field,
                    old_value=old_val,
                    new_value=new_val,
                    ip_address=request.environ.get('REMOTE_ADDR')
                )
                db.session.add(log_entry)
            
            db.session.commit()
            flash('Profil berhasil diperbarui!', 'success')
            return redirect(url_for('profile'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating profile: {str(e)}")
            flash('Terjadi kesalahan saat memperbarui profil!', 'error')
    
    return render_template('edit_profile.html', student=student)

@app.route('/profile/change-password', methods=['GET', 'POST'])
def change_password():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        student = Student.query.get(session['student_id'])
        
        # Validate current password
        if student.password != current_password:
            flash('Password lama tidak benar!', 'error')
            return render_template('change_password.html')
        
        # Validate new password
        if len(new_password) < 6:
            flash('Password baru harus minimal 6 karakter!', 'error')
            return render_template('change_password.html')
        
        if new_password != confirm_password:
            flash('Konfirmasi password tidak cocok!', 'error')
            return render_template('change_password.html')
        
        # Update password
        try:
            # Log password change
            log_entry = ProfileChangeLog(
                student_id=student.id,
                field_changed='password',
                old_value='[HIDDEN]',
                new_value='[HIDDEN]',
                ip_address=request.environ.get('REMOTE_ADDR')
            )
            db.session.add(log_entry)
            
            student.password = new_password
            student.updated_at = datetime.utcnow()
            db.session.commit()
            
            flash('Password berhasil diubah!', 'success')
            return redirect(url_for('profile'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error changing password: {str(e)}")
            flash('Terjadi kesalahan saat mengubah password!', 'error')
    
    return render_template('change_password.html')

@app.route('/profile/upload-photo', methods=['POST'])
def upload_profile_photo():
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    if 'photo' not in request.files:
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        try:
            student = Student.query.get(session['student_id'])
            
            # Generate unique filename
            filename = f"profile_{student.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file.filename.rsplit('.', 1)[1].lower()}"
            
            # Create profile photos directory if it doesn't exist
            profile_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'profiles')
            if not os.path.exists(profile_dir):
                os.makedirs(profile_dir)
            
            # Save new photo
            file_path = os.path.join(profile_dir, filename)
            file.save(file_path)
            
            # Delete old photo if it's not the default
            if student.profile_picture and student.profile_picture != 'default-avatar.png':
                old_path = os.path.join(profile_dir, student.profile_picture)
                if os.path.exists(old_path):
                    os.remove(old_path)
            
            # Update database
            old_picture = student.profile_picture
            student.profile_picture = filename
            student.updated_at = datetime.utcnow()
            
            # Log change
            log_entry = ProfileChangeLog(
                student_id=student.id,
                field_changed='profile_picture',
                old_value=old_picture,
                new_value=filename,
                ip_address=request.environ.get('REMOTE_ADDR')
            )
            db.session.add(log_entry)
            db.session.commit()
            
            return jsonify({
                'success': True, 
                'message': 'Foto profil berhasil diperbarui!',
                'photo_url': url_for('static', filename=f'uploads/profiles/{filename}')
            })
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error uploading profile photo: {str(e)}")
            return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengupload foto'}), 500
    
    return jsonify({'success': False, 'message': 'Format file tidak didukung'}), 400

@app.route('/profile/settings', methods=['GET', 'POST'])
def profile_settings():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    student = Student.query.get(session['student_id'])
    
    if request.method == 'POST':
        try:
            # Update notification settings
            notification_enabled = request.form.get('notification_enabled') == 'on'
            theme_preference = request.form.get('theme_preference', 'light')
            
            # Track changes
            changes = []
            if student.notification_enabled != notification_enabled:
                changes.append(('notification_enabled', str(student.notification_enabled), str(notification_enabled)))
                student.notification_enabled = notification_enabled
            
            if student.theme_preference != theme_preference:
                changes.append(('theme_preference', student.theme_preference, theme_preference))
                student.theme_preference = theme_preference
            
            # Log changes
            for field, old_val, new_val in changes:
                log_entry = ProfileChangeLog(
                    student_id=student.id,
                    field_changed=field,
                    old_value=old_val,
                    new_value=new_val,
                    ip_address=request.environ.get('REMOTE_ADDR')
                )
                db.session.add(log_entry)
            
            student.updated_at = datetime.utcnow()
            db.session.commit()
            
            flash('Pengaturan berhasil disimpan!', 'success')
            return redirect(url_for('profile_settings'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating settings: {str(e)}")
            flash('Terjadi kesalahan saat menyimpan pengaturan!', 'error')
    
    return render_template('profile_settings.html', student=student)

@app.route('/admin/rekap/<int:year>/<int:month>')
def admin_download_rekap(year, month):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    try:
        students = Student.query.order_by(Student.name).all()
        attendances = Attendance.query.filter(
            db.extract('year', Attendance.date) == year,
            db.extract('month', Attendance.date) == month
        ).all()
        attendance_dict = {}
        for attendance in attendances:
            key = (attendance.student_id, attendance.date)
            attendance_dict[key] = attendance.status
        cal = calendar.monthcalendar(year, month)
        days = [day for week in cal for day in week if day != 0]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Rekap Absensi {calendar.month_name[month]} {year}"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        ws.merge_cells('A1:' + get_column_letter(len(days) + 6) + '1')
        ws['A1'] = f"REKAP ABSENSI SISWA PKL - {calendar.month_name[month].upper()} {year}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        headers = ['No', 'Nama Siswa', 'Email'] + [f'{day}' for day in days] + ['Hadir', 'Sakit', 'Izin', 'Tidak Hadir']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        for row_num, student in enumerate(students, 4):
            ws.cell(row=row_num, column=1, value=row_num - 3).border = border
            ws.cell(row=row_num, column=2, value=student.name).border = border
            ws.cell(row=row_num, column=3, value=student.email).border = border
            hadir_count = 0
            sakit_count = 0
            izin_count = 0
            tidak_hadir_count = 0
            for col_num, day in enumerate(days, 4):
                date = datetime(year, month, day).date()
                status = attendance_dict.get((student.id, date), 'Tidak Hadir')
                if status in ['hadir', 'Hadir']:
                    display_status = 'H'
                    hadir_count += 1
                    cell_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif status in ['sakit', 'Sakit']:
                    display_status = 'S'
                    sakit_count += 1
                    cell_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif status in ['izin', 'Izin']:
                    display_status = 'I'
                    izin_count += 1
                    cell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    display_status = 'T'
                    tidak_hadir_count += 1
                    cell_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                cell = ws.cell(row=row_num, column=col_num, value=display_status)
                cell.alignment = center_alignment
                cell.border = border
                cell.fill = cell_fill
            summary_start_col = len(days) + 4
            ws.cell(row=row_num, column=summary_start_col, value=hadir_count).border = border
            ws.cell(row=row_num, column=summary_start_col + 1, value=sakit_count).border = border
            ws.cell(row=row_num, column=summary_start_col + 2, value=izin_count).border = border
            ws.cell(row=row_num, column=summary_start_col + 3, value=tidak_hadir_count).border = border
        legend_row = len(students) + 6
        ws.cell(row=legend_row, column=1, value="KETERANGAN:").font = Font(bold=True)
        ws.cell(row=legend_row + 1, column=1, value="H = Hadir")
        ws.cell(row=legend_row + 2, column=1, value="S = Sakit")
        ws.cell(row=legend_row + 3, column=1, value="I = Izin")
        ws.cell(row=legend_row + 4, column=1, value="T = Tidak Hadir")
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[3].height = 20
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=Rekap_Absensi_{calendar.month_name[month]}_{year}.xlsx'
        return response
    except Exception as e:
        app.logger.error(f"Error generating rekap: {str(e)}")
        flash('Terjadi kesalahan saat membuat rekap!', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/rekap-csv/<int:year>/<int:month>')
def admin_download_rekap_csv(year, month):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    import calendar
    import csv
    students = Student.query.order_by(Student.name).all()
    attendances = Attendance.query.filter(
        db.extract('year', Attendance.date) == year,
        db.extract('month', Attendance.date) == month
    ).all()
    attendance_dict = {}
    for attendance in attendances:
        key = (attendance.student_id, attendance.date)
        attendance_dict[key] = attendance.status
    cal = calendar.monthcalendar(year, month)
    days = [day for week in cal for day in week if day != 0]
    output = io.StringIO()
    writer = csv.writer(output)
    header = ['No', 'Nama Siswa', 'Email'] + [str(day) for day in days] + ['Hadir', 'Sakit', 'Izin', 'Tidak Hadir']
    writer.writerow(header)
    for idx, student in enumerate(students, 1):
        hadir = sakit = izin = tidak_hadir = 0
        row = [idx, student.name, student.email]
        for day in days:
            date = datetime(year, month, day).date()
            status = attendance_dict.get((student.id, date), 'Tidak Hadir')
            if status in ['hadir', 'Hadir']:
                row.append('H')
                hadir += 1
            elif status in ['sakit', 'Sakit']:
                row.append('S')
                sakit += 1
            elif status in ['izin', 'Izin']:
                row.append('I')
                izin += 1
            else:
                row.append('T')
                tidak_hadir += 1
        row += [hadir, sakit, izin, tidak_hadir]
        writer.writerow(row)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
    response.headers['Content-Disposition'] = f'attachment; filename=Rekap_Absensi_{calendar.month_name[month]}_{year}.csv'
    return response

@app.route('/admin/rekap/view/<int:year>/<int:month>')
def admin_view_rekap(year, month):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    import calendar
    students = Student.query.order_by(Student.name).all()
    attendances = Attendance.query.filter(
        db.extract('year', Attendance.date) == year,
        db.extract('month', Attendance.date) == month
    ).all()
    attendance_dict = {}
    for attendance in attendances:
        key = (attendance.student_id, attendance.date)
        attendance_dict[key] = attendance.status
    cal = calendar.monthcalendar(year, month)
    days = [day for week in cal for day in week if day != 0]
    table_data = []
    for student in students:
        statuses = []
        hadir = sakit = izin = tidak_hadir = 0
        for day in days:
            date = datetime(year, month, day).date()
            status = attendance_dict.get((student.id, date), 'Tidak Hadir')
            if status in ['hadir', 'Hadir']:
                css_class = 'hadir'
                hadir += 1
            elif status in ['sakit', 'Sakit']:
                css_class = 'sakit'
                sakit += 1
            elif status in ['izin', 'Izin']:
                css_class = 'izin'
                izin += 1
            else:
                css_class = 'tidak_hadir'
                tidak_hadir += 1
            statuses.append({'status': status.title(), 'css_class': css_class})
        table_data.append({
            'name': student.name,
            'email': student.email,
            'statuses': statuses,
            'summary': {
                'hadir': hadir,
                'sakit': sakit,
                'izin': izin,
                'tidak_hadir': tidak_hadir
            }
        })
    month_name = calendar.month_name[month]
    # Hitung prev/next month & year
    if month == 1:
        prev_month = 12
        prev_year = year - 1
    else:
        prev_month = month - 1
        prev_year = year
    if month == 12:
        next_month = 1
        next_year = year + 1
    else:
        next_month = month + 1
        next_year = year

    return render_template('admin/rekap_print.html',
        year=year,
        month=month,
        month_name=month_name,
        days=days,
        table_data=table_data,
        prev_month=prev_month,
        prev_year=prev_year,
        next_month=next_month,
        next_year=next_year)